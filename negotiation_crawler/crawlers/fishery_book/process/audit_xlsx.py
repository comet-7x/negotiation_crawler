"""Build the teacher-facing XLSX audit table from BookRecords.

Required columns (in order): 类别 / 年份 / 文件名 / 标题 / 下载链接 / 页数 / 大小(KB) / 格式
followed by diagnostic columns. Rows are colour-coded by status so completeness
gaps are visible at a glance.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from ..models import BookRecord, Status

# (header, attribute, width)
COLUMNS: list[tuple[str, str, int]] = [
    ("类别", "category", 18),
    ("年份", "year", 8),
    ("文件名", "filename", 26),
    ("标题", "title", 50),
    ("下载链接", "download_url", 46),
    ("页数", "pages", 8),
    ("大小(KB)", "size_kb", 12),
    ("格式", "fmt", 10),
    # diagnostics
    ("状态", "status", 12),
    ("匹配分", "match_score", 9),
    ("Handle", "handle", 24),
    ("来源", "source", 9),
    ("种子ID", "seed_id", 18),
    ("备注", "note", 40),
]

STATUS_FILL = {
    Status.FOUND: "C6EFCE",      # green
    Status.LEGACY: "DDEBF7",     # light blue
    Status.AMBIGUOUS: "FFEB9C",  # yellow
    Status.NO_PDF: "FCE4D6",     # orange
    Status.MISSING: "FFC7CE",    # red
    Status.ERROR: "FF9999",      # strong red
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_LINK_FONT = Font(color="0563C1", underline="single")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _summary_rows(records: list[BookRecord]) -> list[tuple[str, int]]:
    counts: dict[str, int] = {}
    for r in records:
        counts[r.status.value] = counts.get(r.status.value, 0) + 1
    return sorted(counts.items())


def build_audit_xlsx(records: list[BookRecord], out_path: Path) -> Path:
    wb = Workbook()

    # ---- main sheet ----
    ws = wb.active
    assert ws is not None
    ws.title = "Audit"
    ws.freeze_panes = "A2"

    for col_idx, (header, _attr, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, rec in enumerate(records, start=2):
        fill = STATUS_FILL.get(rec.status)
        for col_idx, (_header, attr, _width) in enumerate(COLUMNS, start=1):
            value = getattr(rec, attr)
            if attr == "status":
                value = rec.status.value
            elif attr == "size_kb" and value is not None:
                value = round(value, 1)
            elif attr == "match_score" and value is not None:
                value = round(value, 1)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(attr in {"title", "note"}))
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            if attr == "download_url" and value:
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=str(value))
                cell.font = _LINK_FONT

    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(records) + 1}"

    # ---- summary sheet ----
    ss = wb.create_sheet("Summary")
    ss.cell(row=1, column=1, value="状态").font = Font(bold=True)
    ss.cell(row=1, column=2, value="数量").font = Font(bold=True)
    for i, (status, n) in enumerate(_summary_rows(records), start=2):
        ss.cell(row=i, column=1, value=status)
        ss.cell(row=i, column=2, value=n)
        try:
            fill = STATUS_FILL.get(Status(status))
        except ValueError:
            fill = None
        if fill:
            ss.cell(row=i, column=1).fill = PatternFill("solid", fgColor=fill)
    total_row = 2 + len(_summary_rows(records))
    ss.cell(row=total_row, column=1, value="总计").font = Font(bold=True)
    ss.cell(row=total_row, column=2, value=len(records)).font = Font(bold=True)
    ss.column_dimensions["A"].width = 16
    ss.column_dimensions["B"].width = 10

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
