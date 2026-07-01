"""
xlsx_builder.py — Build the Excel report from the SQLite manifest.

Sheets (in order):
  1. "All - 全部"                 master sheet, all records
  2. one sheet per doc_type       (23 sheets when all types have data)
  3. "Statistics - 汇总统计"      per-type counts + total + dedup

Sheet names are bilingual: "English - 中文", truncated to ≤ 31 chars.
Types whose combined name exceeds 31 chars use the _SHEET_OVERRIDE table.
"""
from __future__ import annotations

import logging
import sqlite3
from pathlib import Path

log = logging.getLogger("iotc.processor")

# ── category group display order and English labels ────────────────────────────
GROUP_EN: dict[str, str] = {
    "会议报告类": "Meeting Reports",
    "会议文件类": "Meeting Documents",
    "通函类":     "Circulars",
    "提案类":     "Proposals",
    "合规报告类": "Compliance Reports",
    "国家报告类": "National Reports",
    "信息文件类": "Information Papers",
    "参考报告类": "Reference Reports",
    "科学数据类": "Scientific Data",
    "参考文件类": "Reference Documents",
    "出版物类":   "Publications",
    "指南类":     "Guidelines",
    "通用文件类": "General",
    "行政文件类": "Administrative",
    "其他":       "Others",
}
_GROUP_ORDER = list(GROUP_EN.keys())

# ── predefined short sheet names for doc_types whose full name exceeds 31 chars ─
# Format: doc_type (English) → final sheet title (already ≤ 31 chars)
_SHEET_OVERRIDE: dict[str, str] = {
    "Compliance questionnaires":
        "Compliance Q. - 合规问卷",           # 20
    "Final compliance reports":
        "Final Compliance - 最终合规报告",     # 25
    "Letters of Credentials (available upon request)":
        "LoC (On Request) - 授权书（申请）",  # 26
    "Letters of Credentials (Observers)":
        "LoC (Observers) - 授权书（观察员）", # 25
    "Provisionnal compliance reports":
        "Provisional Cmpl. - 临时合规报告",   # 27
    "Reports from other meetings":
        "Other Mtg Reports - 其他会议报告",   # 26
    "Response to feedback letter":
        "Feedback Response - 反馈信回复",      # 25
    "Stock Assessment Input and Output files":
        "Stock Assessment - 种群评估文件",    # 25
    "Summary compliance reports":
        "Summary Compliance - 合规摘要报告",  # 27
}

# ── column definitions ─────────────────────────────────────────────────────────
_COL_DEFS = [
    ("类别",      "doc_type_zh"),
    ("文档类型组", "category_group"),
    ("年份",      "year"),
    ("文件名",    "_fname"),
    ("标题",      "title"),
    ("Reference", "reference"),
    ("会议",      "meeting"),
    ("届次",      "session"),
    ("下载链接",  "pdf_url"),
    ("页数",      "page_count"),
    ("大小(KB)",  "file_size_kb"),
    ("格式",      "_fmt"),
    ("国家",      "country"),
    ("发布日期",  "circulated"),
    ("作者",      "authors"),
    ("状态",      "status"),
]
_HEADERS = [h for h, _ in _COL_DEFS]
_FIELDS  = [f for _, f in _COL_DEFS]


# ── helpers ────────────────────────────────────────────────────────────────────

def _clean(v: object) -> object:
    if not isinstance(v, str):
        return v
    import re
    v = re.sub(r"[\x00-\x08\x0b-\x1f\x7f]", " ", v)
    return v.strip()


def _row_to_values(row: sqlite3.Row) -> list:
    d = dict(row)
    vals = []
    for field in _FIELDS:
        if field == "_fname":
            url = d.get("pdf_url") or ""
            vals.append(_clean(url.rsplit("/", 1)[-1] if url else ""))
        elif field == "_fmt":
            url = d.get("pdf_url") or ""
            ext = url.rsplit(".", 1)[-1].upper() if "." in url else "PDF"
            vals.append(_clean(ext))
        else:
            vals.append(_clean(d.get(field) or ""))
    return vals


def _sheet_name(doc_type: str, doc_type_zh: str) -> str:
    """Return a bilingual sheet title ≤ 31 chars."""
    if doc_type in _SHEET_OVERRIDE:
        return _SHEET_OVERRIDE[doc_type]
    name = f"{doc_type} - {doc_type_zh}"
    return name[:31]


def _group_sort_key(doc_type: str, group: str) -> tuple:
    """Sort key: category-group order first, then doc_type alphabetically."""
    idx = _GROUP_ORDER.index(group) if group in _GROUP_ORDER else len(_GROUP_ORDER)
    return (idx, doc_type.lower())


# ── main entry point ───────────────────────────────────────────────────────────

def build_xlsx(db_path: Path, xlsx_path: Path) -> None:
    """
    Read all English docs from the manifest and write an Excel workbook.

    Sheet order:
      1. "All - 全部"              – every record
      2. one sheet per doc_type    – sorted by category group then name
      3. "Statistics - 汇总统计"   – per-type counts + total + dedup
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    rows = conn.execute(
        "SELECT * FROM docs WHERE language='en' "
        "ORDER BY category_group, doc_type, year, reference"
    ).fetchall()

    if not rows:
        log.warning("No English documents found — xlsx not written.")
        conn.close()
        return

    wb = openpyxl.Workbook()
    active_ws = wb.active
    if active_ws is not None:
        wb.remove(active_ws)

    # ── group by doc_type ──────────────────────────────────────────────────────
    groups: dict[str, list[sqlite3.Row]] = {}
    group_of: dict[str, str] = {}      # doc_type → category_group
    zh_of: dict[str, str]    = {}      # doc_type → doc_type_zh

    for r in rows:
        dt = r["doc_type"] or "其他"
        groups.setdefault(dt, []).append(r)
        if dt not in zh_of:
            zh_of[dt]    = r["doc_type_zh"] or dt
            group_of[dt] = r["category_group"] or "其他"

    # sorted doc_type list: by category group order, then alphabetically
    sorted_types = sorted(
        groups.keys(),
        key=lambda dt: _group_sort_key(dt, group_of.get(dt, "其他")),
    )

    # ── shared styles ──────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F497D")
    center    = Alignment(horizontal="center", vertical="center")
    link_font = Font(color="0563C1", underline="single")
    _URL_COL  = _HEADERS.index("下载链接") + 1

    def _write_data_sheet(ws, sheet_rows: list) -> None:
        ws.append(_HEADERS)
        for cell in ws[1]:
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center

        for row_idx, r in enumerate(sheet_rows, start=2):
            ws.append(_row_to_values(r))
            url = dict(r).get("pdf_url") or ""
            if url:
                cell = ws.cell(row=row_idx, column=_URL_COL)
                cell.hyperlink = url
                cell.value     = url
                cell.font      = link_font

        for col_idx, header in enumerate(_HEADERS, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(header)
            for (cell,) in ws.iter_rows(
                min_row=2, max_row=min(ws.max_row, 200),
                min_col=col_idx, max_col=col_idx,
            ):
                max_len = max(max_len, min(len(str(cell.value or "")), 60))
            ws.column_dimensions[col_letter].width = max_len + 2

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # ── sheet 1: master "All - 全部" ──────────────────────────────────────────
    ws_all = wb.create_sheet("All - 全部")
    _write_data_sheet(ws_all, rows)
    log.info("  sheet 'All - 全部': %d rows", len(rows))

    # ── sheets 2…N: one per doc_type ──────────────────────────────────────────
    for dt in sorted_types:
        title = _sheet_name(dt, zh_of[dt])
        ws = wb.create_sheet(title)
        _write_data_sheet(ws, groups[dt])
        log.info("  sheet '%s': %d rows", title, len(groups[dt]))

    # ── last sheet: statistics ─────────────────────────────────────────────────
    _write_stats_sheet(wb, sorted_types, groups, group_of, zh_of, rows, conn,
                       hdr_font, hdr_fill, center)

    conn.close()
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    log.info("Saved %s (%d total rows, %d doc-type sheets)", xlsx_path, len(rows), len(sorted_types))


# ── statistics sheet ───────────────────────────────────────────────────────────

def _write_stats_sheet(
    wb,
    sorted_types: list[str],
    groups: dict,
    group_of: dict,
    zh_of: dict,
    all_rows: list,
    conn: sqlite3.Connection,
    hdr_font, hdr_fill, center,
) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.create_sheet("Statistics - 汇总统计")

    # column headers
    col_headers = [
        "所属分组 / Category Group",
        "分组英文 / Group (EN)",
        "类别 / Doc Type",
        "英文名 / English Name",
        "数量 / Count",
    ]
    ws.append(col_headers)
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center

    bold       = Font(bold=True)
    right      = Alignment(horizontal="right")
    # alternating group fill colours
    fill_a     = PatternFill("solid", fgColor="EBF3FB")   # light blue
    fill_b     = PatternFill("solid", fgColor="FAFAFA")   # near-white
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    dedup_fill = PatternFill("solid", fgColor="E2EFDA")

    prev_group   = None
    group_colour = fill_a

    for dt in sorted_types:
        grp_zh = group_of.get(dt, "其他")
        grp_en = GROUP_EN.get(grp_zh, grp_zh)
        zh     = zh_of.get(dt, dt)
        count  = len(groups[dt])

        if grp_zh != prev_group:
            group_colour = fill_b if group_colour is fill_a else fill_a
            prev_group = grp_zh

        ws.append([grp_zh, grp_en, zh, dt, count])
        row = ws[ws.max_row]
        for cell in row:
            cell.fill = group_colour
        row[4].alignment = right   # right-align count

    # blank separator
    ws.append([])

    # 合计 / Total
    ws.append(["合计 / Total", "", "", "All records", len(all_rows)])
    total_row = ws[ws.max_row]
    for cell in total_row:
        cell.font = bold
        cell.fill = total_fill
    total_row[4].alignment = right

    # 去重 / Deduplicated
    dedup = conn.execute(
        "SELECT COUNT(DISTINCT sha256) FROM docs "
        "WHERE language='en' AND sha256 IS NOT NULL AND sha256 != ''"
    ).fetchone()[0]
    ws.append(["去重 / Deduplicated", "", "", "Unique files (SHA-256)", dedup])
    dedup_row = ws[ws.max_row]
    for cell in dedup_row:
        cell.font = bold
        cell.fill = dedup_fill
    dedup_row[4].alignment = right

    # column widths
    ws.column_dimensions["A"].width = 28   # 所属分组
    ws.column_dimensions["B"].width = 22   # 分组英文
    ws.column_dimensions["C"].width = 18   # 类别
    ws.column_dimensions["D"].width = 42   # 英文名 (longest: Stock Assessment…)
    ws.column_dimensions["E"].width = 12   # 数量
    ws.freeze_panes = "A2"

    log.info("  sheet 'Statistics - 汇总统计': %d type rows", len(sorted_types))
