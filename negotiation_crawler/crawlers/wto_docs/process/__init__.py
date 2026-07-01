"""Build the metadata index for wto_docs downloaded documents.

Merges per-series detail listings (docs_manifest/detail_*.jsonl) into one
clean table and writes:
  * XLSX   — index.xlsx     (multi-sheet; one "全部" sheet + one per series)
  * CSV    — index.csv      (utf-8-sig; opens in Excel, queryable by DuckDB)
  * SQLite — index.sqlite   (table `documents`; for text-to-SQL / RAG)

Typed columns: year/pages INTEGER, size_kb REAL, downloadable 0/1.
local_path is filled once the PDF has been downloaded by fetch/download.py.
"""

from __future__ import annotations

import csv
import json
import re
import sqlite3
from pathlib import Path

BODY_FOLDER: dict[str, str] = {
    "GFS":   "01_G-FS_渔业补贴委员会",
    "TN":    "02_TN_谈判",
    "WTMIN": "03_WT-MIN_部长会",
    "WTL":   "04_WT-L_法律文本",
    "WTLET": "05_WT-LET_接受书",
    "GSCM":  "06_G-SCM_补贴通报",
    "WTGC":  "07_WT-GC_总理事会",
    "JOBRL": "09_JOB-RL_室文件",
}

SERIES_ZH: dict[str, str] = {
    "GFS":   "G/FS — 渔业补贴委员会",
    "TN":    "TN/RL — 谈判",
    "WTMIN": "WT/MIN — 部长会",
    "WTL":   "WT/L — 法律文本",
    "WTLET": "WT/LET — 接受书",
    "GSCM":  "G/SCM — 补贴通报",
    "WTGC":  "WT/GC — 总理事会",
    "JOBRL": "JOB/RL — 室文件",
}

XLSX_HEADERS = [
    "序号", "系列", "文档号", "标题", "年份", "日期",
    "大小(KB)", "页数", "访问权限", "已下载", "下载链接", "本地路径",
]

COLUMNS = [
    "doc_code", "symbol", "title", "body", "series", "year", "date",
    "size_kb", "pages", "access", "downloadable",
    "subjects", "local_path", "url",
]


def _int(s: str) -> int | None:
    m = re.search(r"\d+", s or "")
    return int(m.group()) if m else None


def _size_kb(s: str) -> float | None:
    m = re.search(r"([\d.,]+)\s*(KB|MB|GB)", s or "", re.I)
    if not m:
        return None
    n = float(m.group(1).replace(",", ""))
    return round(n * {"KB": 1, "MB": 1024, "GB": 1024 * 1024}[m.group(2).upper()], 1)


def _iso(date: str) -> str:
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", date or "")
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else ""


def _safe(symbol: str) -> str:
    name = symbol
    for ch in r'\\/:*?"<>|':
        name = name.replace(ch, "_")
    return name.strip()


def _local_path(body: str, symbol: str, date_iso: str, library_dir: Path) -> str:
    folder = BODY_FOLDER.get(body, body)
    p = library_dir / folder / f"{_safe(symbol)}.pdf"
    return str(p).replace("\\", "/") if p.exists() else ""


def load_rows(manifest_dir: Path, library_dir: Path) -> list[dict]:
    subject_tags: dict[str, list[str]] = {}
    tags_file = manifest_dir / "subject_tags.json"
    if tags_file.exists():
        subject_tags = json.loads(tags_file.read_text(encoding="utf-8"))

    seen: dict[str, dict] = {}
    for jsonl in sorted(manifest_dir.glob("detail_*.jsonl")):
        for line in jsonl.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            r   = json.loads(line)
            key = (r.get("doc_code") or "").strip() or r.get("symbol", f"_row{len(seen)}")
            seen.setdefault(key, r)

    rows = []
    for key, r in seen.items():
        iso  = _iso(r.get("date", ""))
        body = r.get("body", "")
        rows.append({
            "doc_code":    r.get("doc_code", ""),
            "symbol":      r.get("symbol", ""),
            "title":       r.get("title", ""),
            "body":        body,
            "series":      r.get("series", ""),
            "year":        int(iso[:4]) if iso else None,
            "date":        iso,
            "size_kb":     _size_kb(r.get("size", "")),
            "pages":       _int(r.get("pages", "")),
            "access":      "公开" if r.get("downloadable") else "受限",
            "downloadable": 1 if r.get("downloaded") else 0,
            "subjects":    "; ".join(subject_tags.get(key, [])),
            "local_path":  _local_path(body, r.get("symbol", ""), iso, library_dir),
            "url":         r.get("url", ""),
        })

    rows.sort(key=lambda r: (r["body"], r["series"], r["symbol"]))
    return rows


def _write_sheet(ws, rows: list[dict], start_seq: int = 1) -> None:
    """Write XLSX_HEADERS + data rows to an openpyxl worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10)
    body_font   = Font(name="微软雅黑", size=9)

    ws.append(XLSX_HEADERS)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(rows, start_seq):
        ws.append([
            i,
            SERIES_ZH.get(r["body"], r["body"]),
            r["symbol"],
            r["title"],
            r["year"],
            r["date"],
            r["size_kb"],
            r["pages"],
            r["access"],
            "是" if r["downloadable"] else "否",
            r["url"],
            r["local_path"],
        ])

    col_widths = [6, 22, 20, 60, 6, 12, 10, 6, 10, 8, 55, 55]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font      = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_xlsx(rows: list[dict], path: Path) -> None:
    """Write index.xlsx with a "全部" master sheet and per-series sheets."""
    import openpyxl
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws_all = wb.active
    assert ws_all is not None
    ws_all.title = "全部"
    _write_sheet(ws_all, rows)

    # Per-series sheets in canonical order
    bodies = list(dict.fromkeys(r["body"] for r in rows if r["body"]))
    body_order = list(BODY_FOLDER.keys())
    bodies.sort(key=lambda b: body_order.index(b) if b in body_order else 99)
    for body in bodies:
        subset = [r for r in rows if r["body"] == body]
        raw    = str(SERIES_ZH.get(body, body) or "")
        # openpyxl forbids / \ ? * [ ] : in sheet names; max 31 chars
        for ch in r"/\?*[]:'":
            raw = raw.replace(ch, "-")
        title  = raw[:31]
        ws     = wb.create_sheet(title=title)
        _write_sheet(ws, subset)

    wb.save(path)


def write_csv(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)


def write_sqlite(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        path.unlink()
    con = sqlite3.connect(path)
    con.execute("""CREATE TABLE documents (
        doc_code TEXT, symbol TEXT, title TEXT, body TEXT, series TEXT,
        year INTEGER, date TEXT, size_kb REAL, pages INTEGER,
        access TEXT, downloadable INTEGER,
        subjects TEXT, local_path TEXT, url TEXT)""")
    con.executemany(
        f"INSERT INTO documents VALUES ({chr(44).join([chr(63)] * len(COLUMNS))})",
        [[r[c] for c in COLUMNS] for r in rows],
    )
    con.execute("CREATE INDEX idx_body   ON documents(body)")
    con.execute("CREATE INDEX idx_year   ON documents(year)")
    con.execute("CREATE INDEX idx_series ON documents(series)")
    con.commit()
    con.close()


def build_index(
    manifest_dir: Path,
    library_dir: Path,
    out_dir: Path | None = None,
) -> dict:
    """Merge all detail_*.jsonl into index.xlsx + index.csv + index.sqlite.

    Args:
        manifest_dir: directory containing detail_*.jsonl files.
        library_dir:  root of the downloaded PDF library (for local_path lookup).
        out_dir:      where to write outputs; defaults to manifest_dir.

    Returns summary dict: {total, downloadable, restricted, xlsx, csv, sqlite}.
    """
    out = out_dir or manifest_dir
    out.mkdir(parents=True, exist_ok=True)

    rows = load_rows(manifest_dir, library_dir)
    if not rows:
        print("no detail_*.jsonl found — index not written")
        return {"total": 0, "downloadable": 0, "restricted": 0}

    xlsx_path   = out / "index.xlsx"
    csv_path    = out / "index.csv"
    sqlite_path = out / "index.sqlite"

    write_xlsx(rows, xlsx_path)
    write_csv(rows, csv_path)
    write_sqlite(rows, sqlite_path)

    dl = sum(r["downloadable"] for r in rows)
    print(f"index: {len(rows)} rows  downloadable={dl}  restricted={len(rows) - dl}")
    print(f"  XLSX   -> {xlsx_path}")
    print(f"  CSV    -> {csv_path}")
    print(f"  SQLite -> {sqlite_path}  (table: documents)")

    return {
        "total":        len(rows),
        "downloadable": dl,
        "restricted":   len(rows) - dl,
        "xlsx":         str(xlsx_path),
        "csv":          str(csv_path),
        "sqlite":       str(sqlite_path),
    }
