"""xlsx audit export for the wto_site crawl.

Three sheets:
  台账       — full ledger, one URL per row, colour-coded by status
  统计       — counts by status / category / file-type
  边界定义   — boundary/scope rules in plain language

Public entry point:
  build_xlsx(out_dir: Path) -> Path | None
    Reads crawl.db from out_dir, writes audit.xlsx there, returns the path.
    Returns None if the database does not exist.
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import Config
from ..storage.db import DB

# ── Style constants ───────────────────────────────────────────────────────────
_BLUE_FILL  = PatternFill("solid", fgColor="1F4E78")
_GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
_AMBER_FILL = PatternFill("solid", fgColor="FFF2CC")
_RED_FILL   = PatternFill("solid", fgColor="FFDCE1")
_GRAY_FILL  = PatternFill("solid", fgColor="F2F2F2")
_WHITE_FONT = Font(color="FFFFFF", bold=True, name="微软雅黑", size=10)
_BOLD_FONT  = Font(bold=True, name="微软雅黑", size=10)
_CELL_FONT  = Font(name="等线", size=10)

# ── doctype → Chinese label ───────────────────────────────────────────────────
_CATEGORY_CN: dict[str, str] = {
    "legal_text":               "法律文本",
    "international_instrument": "国际文书",
    "publication":              "出版物",
    "ministerial":              "部长级文件",
    "submission":               "谈判提案",
    "meeting_doc":              "会议文件",
    "briefing":                 "简报",
    "news":                     "新闻",
    "navigation":               "参考页面",
    "other":                    "其他",
}

# ── Common filename → Chinese display name ────────────────────────────────────
_CN_NAMES: dict[str, str] = {
    "1969_vclt.pdf":            "维也纳条约法公约（1969）",
    "1982_unclos.pdf":          "联合国海洋法公约（1982）",
    "1995_unfsa.pdf":           "联合国跨界鱼类种群协定（1995）",
    "1995_fao_ccrf.pdf":        "FAO 负责任渔业行为守则（1995）",
    "2001_ipoa_iuu.pdf":        "FAO IUU捕鱼国际行动计划（2001）",
    "2009_psma.pdf":            "港口国措施协定（2009）",
    "2014_vg_fsp.pdf":          "FAO 船旗国表现自愿准则（2014）",
    "2015_fao_ssf.pdf":         "FAO 小规模渔业自愿准则（2015）",
    "2017_vg_cds.pdf":          "FAO 渔获文件方案自愿准则（2017）",
    "fish_factsheet_e.pdf":     "渔业补贴情况说明书",
    "fishagree_e.pdf":          "渔业补贴协定宣传册",
    "implementfishagreement22_e.pdf": "实施渔业补贴协定出版物（完整版）",
    "24-scm.pdf":               "补贴与反补贴措施协定（SCM协定）",
    "WT_MIN22_33.pdf":          "渔业补贴协定全文 WT/MIN22/33（2022）",
    "WT_MIN17_64.pdf":          "布宜诺斯艾利斯部长级决定 WT/MIN17/64（2017）",
    "TN_RL_31.pdf":             "规则谈判组 TN/RL/31",
}


def _file_type(url: str, content_type: str | None) -> str:
    ext = Path(url.split("?")[0]).suffix.lower()
    mapping = {
        ".pdf": "PDF", ".doc": "Word", ".docx": "Word",
        ".xls": "Excel", ".xlsx": "Excel",
        ".ppt": "PPT", ".pptx": "PPT",
        ".zip": "压缩包", ".rar": "压缩包", ".7z": "压缩包",
        ".csv": "CSV", ".rtf": "RTF", ".txt": "文本",
        ".htm": "页面", ".html": "页面", ".aspx": "页面",
    }
    if ext in mapping:
        return mapping[ext]
    ct = (content_type or "").lower()
    if "pdf" in ct:
        return "PDF"
    if "html" in ct:
        return "页面"
    return "文件"


def _cn_name(url: str, local_path: str | None) -> str:
    candidates = []
    if local_path:
        candidates.append(Path(local_path).name)
    candidates.append(url.split("/")[-1].split("?")[0])
    for c in candidates:
        if c in _CN_NAMES:
            return _CN_NAMES[c]
    return ""


def _status_cn(fetch_status: str, scope_action: str) -> str:
    mapping = {
        ("done",    "collect"): "已下载",
        ("done",    "fetch"):   "已爬取（转Markdown）",
        ("skipped", "skip"):    "已跳过",
        ("failed",  "fetch"):   "下载失败",
        ("failed",  "collect"): "下载失败",
        ("pending", "fetch"):   "待爬取",
        ("pending", "collect"): "待下载",
    }
    return mapping.get((fetch_status, scope_action), fetch_status)


def _reason_cn(scope_reason: str | None, fetch_status: str,
               error: str | None) -> str:
    if fetch_status == "failed":
        return f"请求失败：{error or '未知'}"
    if not scope_reason:
        return "已收录" if fetch_status == "done" else ""
    r = scope_reason.lower()
    if "beyond hop budget" in r:
        return f"超出跳数限制（hops>{r.split('>')[-1].strip().rstrip(')')}）"
    if "external host" in r:
        return "外部站点，不在收集范围"
    if "core subtree" in r:
        return "核心子树内，完整收录"
    if "adjacent section" in r:
        return "相邻章节，单跳收录"
    if "file linked" in r:
        return "核心页面链出的文件，自动收录"
    if "extra_seed" in r:
        return "手动补充种子"
    if "seed" in r:
        return "爬取起点（种子）"
    if "non-english" in r:
        return "非英语页面，已跳过"
    return scope_reason


def export(db_path: str, cfg: Config, xlsx_path: str) -> None:
    """Write the 3-sheet audit xlsx from a completed crawl database."""
    db = DB(db_path)
    rows = db.all_rows()
    out_dir = Path(db_path).parent

    # Build parent-page and anchor maps for the ledger sheet
    parent_map: dict[int, str] = {}
    anchor_map: dict[int, str] = {}
    for row in db.conn.execute("""
        SELECT l.to_id, u.url_canonical, l.anchor
        FROM links l JOIN urls u ON l.from_id = u.id
        ORDER BY l.from_id ASC
    """).fetchall():
        tid = row["to_id"]
        if tid not in parent_map:
            parent_map[tid] = row["url_canonical"]
            anchor_map[tid] = row["anchor"] or ""

    wb = Workbook()

    # ── Sheet 1: 台账 ────────────────────────────────────────────────────────
    ws = wb.active
    assert ws is not None
    ws.title = "台账"
    ws.freeze_panes = "A2"

    HEADERS = [
        ("序号",         6),
        ("类型",         8),
        ("类别",        12),
        ("名称（中文）", 30),
        ("英文标题",    36),
        ("父页面",      50),
        ("原始URL",     60),
        ("本地路径",    44),
        ("状态",        14),
        ("原因",        36),
        ("人工审核",    12),
    ]
    for i, (name, width) in enumerate(HEADERS, 1):
        c = ws.cell(1, i, name)
        c.fill = _BLUE_FILL
        c.font = _WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 20

    def sort_key(r):
        status_order = {"done": 0, "pending": 1, "skipped": 2, "failed": 3}
        kind_order   = {"file": 0, "page": 1}
        return (
            status_order.get(r["fetch_status"], 9),
            kind_order.get(r["kind"], 9),
            r["id"],
        )

    for seq, r in enumerate(sorted(rows, key=sort_key), 1):
        status = r["fetch_status"]
        action = r["scope_action"] or ""

        lpath = r["local_path"] or ""
        try:
            lpath_rel = str(Path(lpath).relative_to(out_dir)) if lpath else ""
        except ValueError:
            lpath_rel = lpath

        en_title = (r["title"] or "").strip() or anchor_map.get(r["id"], "")

        row_data = [
            seq,
            _file_type(r["url_canonical"], r["content_type"]),
            _CATEGORY_CN.get(r["doctype"] or "", ""),
            _cn_name(r["url_canonical"], r["local_path"]),
            en_title,
            parent_map.get(r["id"], ""),
            r["url_canonical"],
            lpath_rel,
            _status_cn(status, action),
            _reason_cn(r["scope_reason"], status, r["error"]),
            r["human_verdict"] or "",
        ]
        ws.append(row_data)

        fill = (
            _GREEN_FILL if status == "done"
            else _AMBER_FILL if status == "skipped"
            else _RED_FILL if status == "failed"
            else _GRAY_FILL
        )
        row_idx = ws.max_row
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row_idx, col)
            cell.fill = fill
            cell.font = _CELL_FONT
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    # ── Sheet 2: 统计 ────────────────────────────────────────────────────────
    s = wb.create_sheet("统计")

    by_status  = Counter(r["fetch_status"] for r in rows)
    by_cat     = Counter(
        _CATEGORY_CN.get(r["doctype"] or "", "其他")
        for r in rows if r["fetch_status"] == "done"
    )
    by_type    = Counter(
        _file_type(r["url_canonical"], r["content_type"])
        for r in rows if r["fetch_status"] == "done" and r["kind"] == "file"
    )
    total_bytes = sum(r["size_bytes"] or 0 for r in rows)
    file_done   = sum(1 for r in rows if r["kind"] == "file" and r["fetch_status"] == "done")
    page_done   = sum(1 for r in rows if r["kind"] == "page" and r["fetch_status"] == "done")

    def stat_block(title, pairs, start_row):
        cell = s.cell(start_row, 1, title)
        cell.font = _BOLD_FONT
        cell.fill = PatternFill("solid", fgColor="D6E4F0")
        row = start_row + 1
        for k, v in pairs:
            s.cell(row, 1, k).font = _CELL_FONT
            s.cell(row, 2, v).font = _CELL_FONT
            row += 1
        return row + 1

    nxt = stat_block("抓取状态分布", [
        ("已下载（done）",      by_status.get("done", 0)),
        ("待爬取（pending）",   by_status.get("pending", 0)),
        ("已跳过（skipped）",   by_status.get("skipped", 0)),
        ("失败（failed）",      by_status.get("failed", 0)),
    ], 1)
    nxt = stat_block("已下载文件统计", [
        ("文件数量（文件类）",  file_done),
        ("页面数量（Markdown）", page_done),
        ("总下载量（字节）",    total_bytes),
        ("总下载量（MB）",      round(total_bytes / 1024 / 1024, 2)),
    ], nxt)
    nxt = stat_block("已下载 · 按类别", sorted(by_cat.items()), nxt)
    nxt = stat_block("已下载文件 · 按类型", sorted(by_type.items()), nxt)

    failed = [r for r in rows if r["fetch_status"] == "failed"]
    if failed:
        cell = s.cell(nxt, 1, "失败清单")
        cell.font = _BOLD_FONT
        cell.fill = PatternFill("solid", fgColor="FFDCE1")
        nxt += 1
        for r in failed:
            s.cell(nxt, 1, r["url_canonical"]).font = _CELL_FONT
            s.cell(nxt, 2, r["error"] or "").font = _CELL_FONT
            nxt += 1

    s.column_dimensions["A"].width = 36
    s.column_dimensions["B"].width = 20

    # ── Sheet 3: 边界定义 ────────────────────────────────────────────────────
    b = wb.create_sheet("边界定义")
    for i, h in enumerate(["边界项", "取值", "说明"], 1):
        c = b.cell(1, i, h)
        c.fill = _BLUE_FILL
        c.font = _WHITE_FONT

    defs = [
        ("种子 URL",       cfg.seed_url,
         "主爬取入口，从此页面开始递归"),
        ("补充种子",       "\n".join(cfg.extra_seeds),
         "手动注入 frontier 的补充起点（绕过链接发现限制）"),
        ("核心子树前缀",   cfg.core_prefix,
         "URL path 以此开头 = 核心范围，hops=0，正常递归"),
        ("相邻跳数预算",   cfg.max_hops_outside,
         "允许离开核心前缀的最大步数（1=相邻一跳）"),
        ("页面遍历域",     cfg.page_host_suffix,
         "页面只在此注册域内遍历"),
        ("文件采集域",     cfg.file_host_suffix,
         "文件（PDF/Doc等）可从此域任意子域采集"),
        ("语言限制",       "仅英语（/english/、无 _f/_s 后缀）",
         "路径含 /french/、/spanish/ 或文件名以 _f、_s 结尾的 URL 被跳过"),
        ("文件类型",       ", ".join(cfg.file_extensions),
         "命中即下载、绝不展开"),
        ("linkdoldoc 处理",
         "javascript:linkdoldoc('SYMBOL','') → docs.wto.org 直链",
         "WTO 网站专用 JS 函数，已在 extract.py 中专项解析"),
        ("完整性判据",     "frontier 清空 + 本表人工对账",
         "frontier 空只是必要条件；越界/失败/需渲染清单需人工解释"),
    ]
    for d in defs:
        b.append(list(d))

    b.column_dimensions["A"].width = 20
    b.column_dimensions["B"].width = 52
    b.column_dimensions["C"].width = 44
    b.freeze_panes = "A2"
    b.auto_filter.ref = f"A1:C{b.max_row}"
    for row in b.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = _CELL_FONT
    b.row_dimensions[1].height = 18

    # Optional "需关注" sheet for items needing attention
    needs_render = [r for r in rows if r["needs_render"]]
    if needs_render or failed:
        nr = wb.create_sheet("需关注")
        nr.append(["类别", "URL", "说明"])
        for c in nr[1]:
            c.fill = _BLUE_FILL
            c.font = _WHITE_FONT
        for r in failed:
            nr.append(["下载失败", r["url_canonical"], r["error"] or ""])
        for r in needs_render:
            nr.append(["疑似需JS渲染", r["url_canonical"], "正文极短，脚本多"])
        nr.column_dimensions["A"].width = 14
        nr.column_dimensions["B"].width = 64
        nr.column_dimensions["C"].width = 40

    wb.save(xlsx_path)
    db.close()


def build_xlsx(out_dir: Path) -> Path | None:
    """CLI entry point: read crawl.db from out_dir, write audit.xlsx.

    Returns the xlsx path, or None if the database does not exist.
    """
    db_path = out_dir / "crawl.db"
    if not db_path.exists():
        print("wto_site: crawl.db not found — xlsx not written")
        return None

    xlsx_path = out_dir / "audit.xlsx"
    cfg = Config(out_dir=str(out_dir))
    export(str(db_path), cfg, str(xlsx_path))

    # Print a quick summary
    db = DB(str(db_path))
    c = db.counts()
    db.close()
    print(f"wto_site audit.xlsx: {c['total']} URLs  done={c['done']}")
    print(f"  XLSX -> {xlsx_path}")
    return xlsx_path
