"""
unified_db.py — 统一文档表：单宽表供 text-to-SQL 智能体查询。

表结构设计原则
--------------
1. 单宽表 documents —— text-to-SQL 无需 JOIN，直接过滤 source 列
2. 公有字段：4 个模块都可能填写（NULL 表示该模块暂无此数据）
3. 特殊字段：仅某模块有意义，COMMENT 中标注 [来源]
4. 每列均有 COMMENT，供 LLM 理解字段语义
5. source_id = 各模块的自然主键（pdf_url / doc_code / source_url）

来源标识（source 列）
---------------------
  iotc          印度洋金枪鱼委员会文件库
  wto_docs      WTO Documents Online 正式文件
  wto_site      WTO 渔业补贴协议主网站及子页面
  fishery_book  FAO 渔业出版物（DSpace 知识库）

公有字段
--------
title, category, year, pub_date, status,
filename, file_format, file_size_kb, page_count,
download_url, local_path, is_downloadable, is_downloaded

特殊字段（按来源分组，在 SQL DDL 中以 COMMENT 标注）
------------------------------------------------------
IOTC:         category_group, reference, meeting, session, country, authors
WTO Docs:     series, doc_number, subject, doc_code, institution
WTO Site:     content_type, title_zh, parent_url, source_url
Fishery Book: match_score, handle, book_source, seed_id, notes
"""
from __future__ import annotations

import re
import sqlite3
from datetime import date
from pathlib import Path
from typing import Any

# ── DDL ───────────────────────────────────────────────────────────────────────

#: MySQL DDL.  Call create_table_sqlite() for SQLite (strips MySQL-isms).
MYSQL_DDL = """\
CREATE TABLE IF NOT EXISTS documents (
    id              BIGINT        AUTO_INCREMENT PRIMARY KEY,

    -- ── 来源标识 ─────────────────────────────────────────────────────────────
    source          VARCHAR(20)   NOT NULL
                    COMMENT '数据来源: iotc / wto_docs / wto_site / fishery_book',
    source_id       VARCHAR(512)
                    COMMENT '各来源的自然主键：IOTC=pdf_url, WTO文件=doc_code, WTO网站=source_url, FAO书籍=handle',

    -- ── 公有字段：标题与分类 ─────────────────────────────────────────────────
    title           TEXT
                    COMMENT '标题（英文）',
    title_zh        TEXT
                    COMMENT '中文标题或页面名称 [wto_site]',
    category        VARCHAR(150)
                    COMMENT '文档类别，各模块含义不同：IOTC=doc_type_zh, wto_site=类别, fishery_book=category',
    category_group  VARCHAR(150)
                    COMMENT '上层分类组 [iotc=文档类型组; wto_docs=系列如G/FS]',

    -- ── 公有字段：时间 ───────────────────────────────────────────────────────
    year            SMALLINT
                    COMMENT '出版/发布年份（整数）',
    pub_date        DATE
                    COMMENT '精确发布日期 [iotc=circulated; wto_docs=日期]',

    -- ── 公有字段：文件属性 ───────────────────────────────────────────────────
    filename        VARCHAR(500)
                    COMMENT '原始文件名（含扩展名）',
    file_format     VARCHAR(20)
                    COMMENT '文件格式：PDF / DOCX / HTML / ...',
    file_size_kb    FLOAT
                    COMMENT '文件大小（KB）',
    page_count      INT
                    COMMENT '页数',

    -- ── 公有字段：获取状态 ───────────────────────────────────────────────────
    download_url    TEXT
                    COMMENT '下载/原始链接',
    local_path      TEXT
                    COMMENT '本地存储路径（相对于 output_dir）',
    is_downloadable TINYINT(1)
                    COMMENT '是否可下载（1=是 0=否/受限）',
    is_downloaded   TINYINT(1)
                    COMMENT '是否已下载（1=是 0=否）',
    status          VARCHAR(50)
                    COMMENT '状态：downloaded / pending / failed / restricted 等',

    -- ── IOTC 特殊字段 ────────────────────────────────────────────────────────
    reference       VARCHAR(200)
                    COMMENT '[iotc] Reference 编号，如 IOTC-2023-SC26-R01',
    meeting         VARCHAR(200)
                    COMMENT '[iotc] 会议名称，如 26th Session of the Scientific Committee',
    session         VARCHAR(50)
                    COMMENT '[iotc] 届次，如 26',
    country         VARCHAR(100)
                    COMMENT '[iotc] 国家（仅国家报告类有值）',
    authors         TEXT
                    COMMENT '[iotc] 作者',

    -- ── WTO 文件库特殊字段 ───────────────────────────────────────────────────
    institution     VARCHAR(200)
                    COMMENT '[wto_docs] 发文机构',
    series          VARCHAR(100)
                    COMMENT '[wto_docs] 文件系列，如 G/FS、TN/RL、WT/MIN',
    doc_number      VARCHAR(200)
                    COMMENT '[wto_docs] 文档号（symbol），如 G/FS/W/123',
    subject         TEXT
                    COMMENT '[wto_docs] 文件主题/摘要',
    doc_code        VARCHAR(200)
                    COMMENT '[wto_docs] 系统内部 Doc 编码',

    -- ── WTO 网站特殊字段 ─────────────────────────────────────────────────────
    content_type    VARCHAR(50)
                    COMMENT '[wto_site] 内容类型：html / pdf / doc / markdown',
    parent_url      TEXT
                    COMMENT '[wto_site] 父页面 URL（爬取来源页）',
    source_url      TEXT
                    COMMENT '[wto_site] 原始 URL',

    -- ── FAO Fishery Book 特殊字段 ────────────────────────────────────────────
    match_score     FLOAT
                    COMMENT '[fishery_book] DSpace 搜索匹配分',
    handle          VARCHAR(200)
                    COMMENT '[fishery_book] DSpace Handle，如 10.4060/cb1234en',
    book_source     VARCHAR(50)
                    COMMENT '[fishery_book] 爬取来源：dspace / legacy',
    seed_id         VARCHAR(100)
                    COMMENT '[fishery_book] 种子集合 ID',
    notes           TEXT
                    COMMENT '[fishery_book] 备注（匹配歧义、错误信息等）',

    -- ── 时间戳 ───────────────────────────────────────────────────────────────
    created_at      DATETIME      DEFAULT CURRENT_TIMESTAMP,
    updated_at      DATETIME      DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,

    UNIQUE KEY  uk_source_id (source, source_id(255)),
    INDEX       idx_source   (source),
    INDEX       idx_category (category),
    INDEX       idx_year     (year),
    INDEX       idx_series   (series),
    INDEX       idx_status   (status),
    INDEX       idx_country  (country),
    FULLTEXT    ft_title     (title)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
  COMMENT='统一文档库，供 text-to-SQL 智能体跨模块查询';
"""

# SQLite version（去掉 MySQL 方言）
_SQLITE_DDL = re.sub(
    r"(AUTO_INCREMENT|ENGINE=\S+|DEFAULT CHARSET=\S+|COLLATE=\S+|ON UPDATE CURRENT_TIMESTAMP)",
    "",
    MYSQL_DDL,
)
_SQLITE_DDL = re.sub(r"TINYINT\(1\)", "INTEGER", _SQLITE_DDL)
_SQLITE_DDL = re.sub(r"BIGINT", "INTEGER", _SQLITE_DDL)
_SQLITE_DDL = re.sub(r"FULLTEXT\s+ft_title\s*\([^)]+\)\s*,?", "", _SQLITE_DDL)
_SQLITE_DDL = re.sub(r",\s*\)", "\n)", _SQLITE_DDL)   # trailing comma cleanup


def create_table_sqlite(conn: sqlite3.Connection) -> None:
    conn.executescript(_SQLITE_DDL)
    conn.commit()


def create_table_mysql(conn) -> None:
    """conn: any PEP-249 MySQL connection (mysql-connector-python, pymysql, etc.)."""
    cur = conn.cursor()
    cur.execute(MYSQL_DDL)
    conn.commit()


# ── 工具函数 ──────────────────────────────────────────────────────────────────

def _str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _int(v: Any) -> int | None:
    if v is None:
        return None
    try:
        return int(str(v).strip().replace(",", ""))
    except (ValueError, TypeError):
        m = re.search(r"\d+", str(v))
        return int(m.group()) if m else None


def _float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).strip().replace(",", ""))
    except (ValueError, TypeError):
        return None


def _bool(v: Any) -> int | None:
    """Convert '是/否', '1/0', True/False, 'Y/N', '✓/✗' → 1/0/None."""
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in {"1", "是", "yes", "y", "true", "✓", "√", "已下载", "可下载"}:
        return 1
    if s in {"0", "否", "no", "n", "false", "✗", "×", "受限", "未下载"}:
        return 0
    return None


def _date(v: Any) -> str | None:
    """Return ISO date string or None."""
    if v is None:
        return None
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    # try common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # fallback: extract 4-digit year only
    m = re.search(r"\d{4}", s)
    return f"{m.group()}-01-01" if m else None


def _read_xlsx_sheet(xlsx_path: Path, sheet_name: str | None = None) -> tuple[list[str], list[list]]:
    """Return (headers, data_rows) from the first matching sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = None
    if sheet_name:
        for name in wb.sheetnames:
            if sheet_name in name:
                ws = wb[name]
                break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data    = [list(r) for r in rows[1:] if any(c is not None for c in r)]
    return headers, data


def _upsert(conn: sqlite3.Connection, row: dict[str, Any]) -> None:
    cols   = list(row.keys())
    vals   = list(row.values())
    ph     = ", ".join("?" * len(cols))
    col_s  = ", ".join(cols)
    update = ", ".join(f"{c}=excluded.{c}" for c in cols if c not in ("source", "source_id"))
    sql = (
        f"INSERT INTO documents ({col_s}) VALUES ({ph}) "
        f"ON CONFLICT(source, source_id) DO UPDATE SET {update}"
    )
    conn.execute(sql, vals)


# ── IOTC ─────────────────────────────────────────────────────────────────────

# xlsx 列名 → DB 字段映射
_IOTC_COL_MAP = {
    "类别":      "category",
    "文档类型组": "category_group",
    "年份":      "year",
    "文件名":    "filename",
    "标题":      "title",
    "Reference": "reference",
    "会议":      "meeting",
    "届次":      "session",
    "下载链接":  "download_url",
    "页数":      "page_count",
    "大小(KB)":  "file_size_kb",
    "格式":      "file_format",
    "国家":      "country",
    "发布日期":  "pub_date",
    "作者":      "authors",
    "状态":      "status",
}


def from_iotc_xlsx(xlsx_path: Path, conn: sqlite3.Connection) -> int:
    """
    读取 IOTC index.xlsx（All-全部 sheet）→ 写入 documents 表。
    返回插入/更新行数。
    """
    headers, data = _read_xlsx_sheet(xlsx_path, "All")
    if not headers:
        return 0

    idx = {h: i for i, h in enumerate(headers)}
    n = 0
    for raw in data:
        def _get(col: str):
            i = idx.get(col)
            return raw[i] if i is not None and i < len(raw) else None

        url = _str(_get("下载链接"))
        row: dict[str, Any] = {
            "source":          "iotc",
            "source_id":       url or _str(_get("Reference")) or _str(_get("标题")),
            "title":           _str(_get("标题")),
            "category":        _str(_get("类别")),
            "category_group":  _str(_get("文档类型组")),
            "year":            _int(_get("年份")),
            "pub_date":        _date(_get("发布日期")),
            "filename":        _str(_get("文件名")),
            "file_format":     _str(_get("格式")),
            "file_size_kb":    _float(_get("大小(KB)")),
            "page_count":      _int(_get("页数")),
            "download_url":    url,
            "status":          _str(_get("状态")),
            "reference":       _str(_get("Reference")),
            "meeting":         _str(_get("会议")),
            "session":         _str(_get("届次")),
            "country":         _str(_get("国家")),
            "authors":         _str(_get("作者")),
        }
        if row["source_id"] is None:
            continue
        _upsert(conn, row)
        n += 1
    conn.commit()
    return n


# ── WTO Docs ─────────────────────────────────────────────────────────────────

def from_wto_docs_xlsx(xlsx_path: Path, conn: sqlite3.Connection) -> int:
    """
    读取 WTO Docs index.xlsx（全部 sheet）→ 写入 documents 表。

    xlsx 现有列：序号, 系列, 文档号, 标题, 年份, 日期, 大小(KB), 页数,
                 访问权限, 已下载, 下载链接, 本地路径
    缺失字段（需从 JSONL/SQLite 补充）：机构, 主题, Doc编码
    """
    headers, data = _read_xlsx_sheet(xlsx_path, "全部")
    if not headers:
        return 0

    idx = {h: i for i, h in enumerate(headers)}
    n = 0
    for raw in data:
        def _get(col: str):
            i = idx.get(col)
            return raw[i] if i is not None and i < len(raw) else None

        url = _str(_get("下载链接"))
        doc_num = _str(_get("文档号"))
        row: dict[str, Any] = {
            "source":          "wto_docs",
            "source_id":       doc_num or url or _str(_get("序号")),
            "title":           _str(_get("标题")),
            "category_group":  _str(_get("系列")),   # series 同时进 category_group
            "series":          _str(_get("系列")),
            "doc_number":      doc_num,
            "year":            _int(_get("年份")),
            "pub_date":        _date(_get("日期")),
            "file_size_kb":    _float(_get("大小(KB)")),
            "page_count":      _int(_get("页数")),
            "is_downloadable": _bool(_get("访问权限")),
            "is_downloaded":   _bool(_get("已下载")),
            "download_url":    url,
            "local_path":      _str(_get("本地路径")),
        }
        if row["source_id"] is None:
            continue
        _upsert(conn, row)
        n += 1
    conn.commit()
    return n


# ── WTO Site ─────────────────────────────────────────────────────────────────

def from_wto_site_xlsx(xlsx_path: Path, conn: sqlite3.Connection) -> int:
    """
    读取 WTO Site index.xlsx（全部 sheet）→ 写入 documents 表。

    xlsx 现有列：序号, 类别, 标题, 类型, 状态, 本地路径, 来源URL
    缺失字段（需从 manifest.jsonl 补充）：title_zh(名称中文), parent_url(父页面)
    """
    headers, data = _read_xlsx_sheet(xlsx_path, "全部")
    if not headers:
        return 0

    idx = {h: i for i, h in enumerate(headers)}
    n = 0
    for raw in data:
        def _get(col: str):
            i = idx.get(col)
            return raw[i] if i is not None and i < len(raw) else None

        src_url = _str(_get("来源URL")) or _str(_get("原始URL")) or _str(_get("source_url"))
        row: dict[str, Any] = {
            "source":        "wto_site",
            "source_id":     src_url or _str(_get("本地路径")),
            "title":         _str(_get("标题")) or _str(_get("英文标题")),
            "title_zh":      _str(_get("名称(中文)")),
            "category":      _str(_get("类别")),
            "content_type":  _str(_get("类型")),
            "status":        _str(_get("状态")),
            "local_path":    _str(_get("本地路径")),
            "source_url":    src_url,
            "parent_url":    _str(_get("父页面")),
        }
        if row["source_id"] is None:
            continue
        _upsert(conn, row)
        n += 1
    conn.commit()
    return n


# ── Fishery Book ──────────────────────────────────────────────────────────────

def from_fishery_book_xlsx(xlsx_path: Path, conn: sqlite3.Connection) -> int:
    """
    读取 fishery_book index.xlsx（Audit sheet）→ 写入 documents 表。

    xlsx 列：类别, 年份, 文件名, 标题, 下载链接, 页数, 大小(KB), 格式,
             状态, 匹配分, Handle, 来源, 种子ID, 备注
    """
    # fishery_book 用 "Audit" sheet（也试 "全部"）
    headers, data = _read_xlsx_sheet(xlsx_path, "Audit")
    if not headers:
        headers, data = _read_xlsx_sheet(xlsx_path, "全部")
    if not headers:
        headers, data = _read_xlsx_sheet(xlsx_path, None)  # first sheet
    if not headers:
        return 0

    idx = {h: i for i, h in enumerate(headers)}
    n = 0
    for raw in data:
        def _get(col: str):
            i = idx.get(col)
            return raw[i] if i is not None and i < len(raw) else None

        handle = _str(_get("Handle"))
        url    = _str(_get("下载链接"))
        row: dict[str, Any] = {
            "source":        "fishery_book",
            "source_id":     handle or url or _str(_get("文件名")),
            "title":         _str(_get("标题")),
            "category":      _str(_get("类别")),
            "year":          _int(_get("年份")),
            "filename":      _str(_get("文件名")),
            "download_url":  url,
            "page_count":    _int(_get("页数")),
            "file_size_kb":  _float(_get("大小(KB)")),
            "file_format":   _str(_get("格式")),
            "status":        _str(_get("状态")),
            "match_score":   _float(_get("匹配分")),
            "handle":        handle,
            "book_source":   _str(_get("来源")),
            "seed_id":       _str(_get("种子ID")),
            "notes":         _str(_get("备注")),
        }
        if row["source_id"] is None:
            continue
        _upsert(conn, row)
        n += 1
    conn.commit()
    return n


# ── 批量导入入口 ──────────────────────────────────────────────────────────────

_IMPORTERS = {
    "iotc":         from_iotc_xlsx,
    "wto_docs":     from_wto_docs_xlsx,
    "wto_site":     from_wto_site_xlsx,
    "fishery_book": from_fishery_book_xlsx,
}


def import_all(
    base_dir: Path,
    db_path: Path,
    *,
    modules: list[str] | None = None,
) -> dict[str, int]:
    """
    从 base_dir/{module}/index.xlsx 批量导入所有模块到 db_path。

    base_dir 结构（与 `run all` 输出一致）：
        base_dir/
          iotc/index.xlsx
          wto_docs/index.xlsx
          wto_site/index.xlsx
          fishery_book/index.xlsx

    返回 {module: rows_inserted} 字典。
    """
    conn = sqlite3.connect(db_path)
    create_table_sqlite(conn)

    results: dict[str, int] = {}
    targets = modules or list(_IMPORTERS.keys())

    for mod in targets:
        xlsx = base_dir / mod / "index.xlsx"
        if not xlsx.exists():
            results[mod] = 0
            continue
        fn = _IMPORTERS[mod]
        n  = fn(xlsx, conn)
        results[mod] = n

    conn.close()
    return results
